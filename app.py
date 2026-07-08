from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash
import pandas as pd
import numpy as np
import sqlite3, os, io, json
from datetime import datetime, date, timedelta
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, mean_squared_error
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import warnings
warnings.filterwarnings('ignore')

app = Flask(__name__)
app.secret_key = 'ayamserayu2025secretkey'
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
DB_PATH = 'ayam_serayu.db'

PRODUK_MAKANAN = [
    # Ayam goreng serayu
    'Ayam goreng serayu',
    'Ayam goreng serayu - Paha',
    'Ayam goreng serayu - Dada',
    # Ayam goreng kremes
    'Ayam goreng kremes',
    'Ayam goreng kremes - Paha',
    'Ayam goreng kremes - Dada',
    # Nasi
    'Nasi',
    'Nasi 1/2',
    'Nasi Uduk.',
    'Nasi uduk 1/2',
    # Paket nasi + ayam
    'Nasi+ayam goreng serayu',
    'Nasi+ayam goreng serayu - Paha',
    'Nasi+ayam goreng serayu - Dada',
    # Sayuran
    'Cah kangkung',
    'Kangkung crispy',
    'Kangkung goreng',
    'Jukut goreng',
    'Pete goreng',
    'Tumis pete',
    # Lauk & lainnya
    'Kulit cabe garam',
    'Kulit goreng',
    'Usus cabe garam',
    'Usus goreng',
    'Tahu',
    'Tempe',
    'Telor crispy',
    'Nila goreng',
    'ati ampela',
    'Sambal',
]
PRODUK_MINUMAN = [
    'Teh manis dingin',
    'Teh manis panas',
    'Teh tawar dingin',
    'Teh tawar panas',
    'Airmineral',
    'Airmineral - Dingin',
    'Airmineral - Biasa',
    'Air putih es',
    'Ait putih',
    'Es batu',
]

FEATURES = [
    'hari_ke', 'hari_dalam_minggu', 'hari_dalam_bulan',
    'bulan', 'minggu_ke',
    'lag_1', 'lag_2', 'lag_3', 'lag_7',
    'rata_rata_3hari', 'rata_rata_7hari'
]

# Fitur untuk prediksi per-item: hilangkan hari_ke dan bulan karena keduanya
# menyebabkan LR mengekstrapolasi tren jangka panjang yang tidak realistis.
# Cukup pola hari-dalam-minggu + lag + rata-rata untuk prediksi 1 hari ke depan.
FEATURES_ITEM = [
    'hari_dalam_minggu', 'hari_dalam_bulan', 'minggu_ke',
    'lag_1', 'lag_2', 'lag_3', 'lag_7',
    'rata_rata_3hari', 'rata_rata_7hari'
]

# ─────────────── DATABASE ───────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS penjualan (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            tanggal   TEXT NOT NULL,
            produk    TEXT NOT NULL,
            kategori  TEXT DEFAULT 'Makanan',
            jumlah    INTEGER NOT NULL DEFAULT 0,
            harga     INTEGER DEFAULT 0
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS prediksi_hasil (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at  TEXT,
            result_json TEXT
        )
    ''')
    for col, typedef in [('kategori', "TEXT DEFAULT 'Makanan'"), ('harga', 'INTEGER DEFAULT 0')]:
        try:
            conn.execute(f'ALTER TABLE penjualan ADD COLUMN {col} {typedef}')
        except:
            pass
    conn.commit()
    conn.close()

init_db()

# ─────────────── HELPERS ───────────────
def clean_number(val):
    if isinstance(val, str):
        val = val.replace('"', '').replace(',', '').strip()
    try:
        return int(float(val))
    except:
        return 0

def detect_columns(df):
    col_map = {}
    for c in df.columns:
        cl = c.lower().replace(' ', '_').replace('&', '_').replace('/', '_')
        cl = '_'.join(filter(None, cl.split('_')))
        if 'tanggal' in cl or 'waktu' in cl or 'date' in cl or 'time' in cl:
            col_map.setdefault('tanggal', c)
        if ('jumlah' in cl and 'produk' in cl) or cl in ('qty', 'quantity', 'jumlah'):
            col_map.setdefault('jumlah_produk', c)
        if 'nama' in cl and ('produk' in cl or 'item' in cl or 'menu' in cl):
            col_map.setdefault('nama_produk', c)
        if 'kategori' in cl or 'category' in cl:
            col_map.setdefault('kategori', c)
        if 'harga' in cl and 'produk' in cl:
            col_map.setdefault('harga_produk', c)
        if 'kotor' in cl or 'gross' in cl:
            col_map.setdefault('penjualan_kotor', c)
    return col_map

def make_features(daily):
    daily = daily.copy().sort_values('tanggal').reset_index(drop=True)
    daily['hari_ke']           = range(1, len(daily) + 1)
    daily['hari_dalam_minggu'] = pd.to_datetime(daily['tanggal']).dt.dayofweek
    daily['hari_dalam_bulan']  = pd.to_datetime(daily['tanggal']).dt.day
    daily['bulan']             = pd.to_datetime(daily['tanggal']).dt.month
    daily['minggu_ke']         = pd.to_datetime(daily['tanggal']).dt.isocalendar().week.astype(int)
    for lag in [1, 2, 3, 7]:
        daily[f'lag_{lag}'] = daily['jumlah_terjual'].shift(lag)
    daily['rata_rata_3hari'] = daily['jumlah_terjual'].shift(1).rolling(3).mean()
    daily['rata_rata_7hari'] = daily['jumlah_terjual'].shift(1).rolling(7).mean()
    return daily.dropna(subset=FEATURES + ['jumlah_terjual']).reset_index(drop=True)

def train_eval(daily, test_size, n_estimators):
    # Gunakan chronological split: data lebih awal = train, data terbaru = test.
    # Ini lebih realistis untuk time series daripada random split,
    # karena model hanya "melihat masa lalu" saat dievaluasi.
    n_test = max(1, int(len(daily) * test_size))
    n_train = len(daily) - n_test

    train = daily.iloc[:n_train]
    test  = daily.iloc[n_train:]

    X, y = daily[FEATURES], daily['jumlah_terjual']
    Xtr, ytr = train[FEATURES], train['jumlah_terjual']
    Xte, yte = test[FEATURES],  test['jumlah_terjual']

    lr = LinearRegression()
    lr.fit(Xtr, ytr)
    plr = lr.predict(Xte)

    rf = RandomForestRegressor(n_estimators=n_estimators, random_state=42)
    rf.fit(Xtr, ytr)
    prf = rf.predict(Xte)

    mae_lr  = round(float(mean_absolute_error(yte, plr)), 2)
    rmse_lr = round(float(np.sqrt(mean_squared_error(yte, plr))), 2)
    mae_rf  = round(float(mean_absolute_error(yte, prf)), 2)
    rmse_rf = round(float(np.sqrt(mean_squared_error(yte, prf))), 2)
    best    = 'Linear Regression' if rmse_lr <= rmse_rf else 'Random Forest'

    return {
        'y_test':        [round(v) for v in yte.tolist()],
        'y_pred_lr':     [round(v) for v in plr.tolist()],
        'y_pred_rf':     [round(v) for v in prf.tolist()],
        'tanggal_test':  [str(t) for t in test['tanggal'].tolist()],
        'mae_lr': mae_lr, 'rmse_lr': rmse_lr,
        'mae_rf': mae_rf, 'rmse_rf': rmse_rf,
        'best_model': best,
        'aktual_list':   daily['jumlah_terjual'].tolist(),
        'tanggal_list':  [str(t) for t in daily['tanggal'].tolist()],
    }

def build_daily_per_produk_from_db(produk):
    conn = get_db()
    rows = conn.execute(
        'SELECT tanggal, kategori, SUM(jumlah) as total FROM penjualan '
        'WHERE produk=? GROUP BY tanggal ORDER BY tanggal',
        (produk,)
    ).fetchall()
    conn.close()
    if not rows:
        return None, None
    df = pd.DataFrame([{
        'tanggal': r['tanggal'],
        'jumlah_terjual': r['total'],
        'kategori': r['kategori'],
    } for r in rows])
    df['tanggal'] = pd.to_datetime(df['tanggal']).dt.date
    kategori = df['kategori'].iloc[-1]
    return df[['tanggal', 'jumlah_terjual']], kategori


def predict_item_horizon(daily_raw, horizon=1, n_estimators=100, buffer_pct=15):
    """
    Prediksi stok untuk horizon hari ke depan secara rekursif.
    horizon=1  → besok
    horizon=7  → 1 minggu (total 7 hari)
    horizon=30 → 1 bulan  (total 30 hari)
    """
    daily = daily_raw.copy().sort_values('tanggal').reset_index(drop=True)
    n = len(daily)
    vals = daily['jumlah_terjual'].astype(float).values.tolist()
    last_date = pd.to_datetime(daily['tanggal'].iloc[-1])

    terjual_terakhir = int(vals[-1])
    rata_7hari = round(float(np.mean(vals[-7:])) if n >= 7 else float(np.mean(vals)), 1)
    cap_max = round(float(np.max(vals[-30:])) * 2) if n >= 5 else round(float(np.max(vals)) * 2)

    rf_model = lr_model = None
    metode = 'Rata-rata'

    if n >= 10:
        try:
            feat_df = make_features(daily.copy())
            if len(feat_df) >= 5:
                X = feat_df[FEATURES_ITEM]
                y = feat_df['jumlah_terjual']
                lr_model = LinearRegression();                                      lr_model.fit(X, y)
                rf_model = RandomForestRegressor(n_estimators=n_estimators, random_state=42); rf_model.fit(X, y)
                metode = 'ML'
        except Exception as e:
            print(f'predict_item_horizon error ({e})')

    def _predict_one(cur_vals, pred_date):
        n_c = len(cur_vals)
        if rf_model:
            feat = pd.DataFrame([{
                'hari_dalam_minggu': pred_date.weekday(),
                'hari_dalam_bulan':  pred_date.day,
                'minggu_ke':         int(pred_date.isocalendar()[1]),
                'lag_1':             float(cur_vals[-1]),
                'lag_2':             float(cur_vals[-2]) if n_c >= 2 else float(cur_vals[-1]),
                'lag_3':             float(cur_vals[-3]) if n_c >= 3 else float(cur_vals[-1]),
                'lag_7':             float(cur_vals[-7]) if n_c >= 7 else float(np.mean(cur_vals)),
                'rata_rata_3hari':   float(np.mean(cur_vals[-3:])) if n_c >= 3 else float(np.mean(cur_vals)),
                'rata_rata_7hari':   float(np.mean(cur_vals[-7:])) if n_c >= 7 else float(np.mean(cur_vals)),
            }])
            p_rf = max(0, min(round(float(rf_model.predict(feat)[0])), cap_max))
            p_lr = max(0, min(round(float(lr_model.predict(feat)[0])), cap_max))
        else:
            avg  = round(float(np.mean(cur_vals[-7:])) if len(cur_vals) >= 7 else float(np.mean(cur_vals)))
            p_rf = p_lr = avg
        return p_rf, p_lr

    cur_vals   = vals.copy()
    preds_rf   = []
    preds_lr   = []

    for step in range(1, horizon + 1):
        pred_date = last_date + pd.Timedelta(days=step)
        p_rf, p_lr = _predict_one(cur_vals, pred_date)
        preds_rf.append(p_rf)
        preds_lr.append(p_lr)
        cur_vals.append(float(p_rf))   # RF sebagai dasar prediksi rekursif

    total_rf   = sum(preds_rf)
    total_lr   = sum(preds_lr)
    prediksi   = preds_rf[0] if horizon == 1 else total_rf
    stok_anjuran = round(prediksi * (1 + buffer_pct / 100))

    return {
        'terjual_terakhir': terjual_terakhir,
        'rata_7hari':       rata_7hari,
        'pred_lr':          preds_lr[0] if horizon == 1 else total_lr,
        'pred_rf':          preds_rf[0] if horizon == 1 else total_rf,
        'prediksi':         prediksi,
        'avg_per_hari':     round(total_rf / horizon, 1),
        'stok_anjuran':     stok_anjuran,
        'metode':           metode,
        'last_date':        str(last_date.date()),
        'next_date':        str((last_date + pd.Timedelta(days=1)).date()),
        'end_date':         str((last_date + pd.Timedelta(days=horizon)).date()),
        'total_data':       n,
        'horizon':          horizon,
    }


def predict_next_day_item(daily_raw, n_estimators=100, buffer_pct=15):
    return predict_item_horizon(daily_raw, horizon=1,
                                n_estimators=n_estimators, buffer_pct=buffer_pct)


def build_daily_from_db(kategori):
    conn = get_db()
    rows = conn.execute(
        'SELECT tanggal, SUM(jumlah) as total FROM penjualan '
        'WHERE LOWER(kategori)=? GROUP BY tanggal ORDER BY tanggal',
        (kategori.lower(),)
    ).fetchall()
    conn.close()
    if len(rows) < 10:
        return None
    df = pd.DataFrame([{'tanggal': r['tanggal'], 'jumlah_terjual': r['total']} for r in rows])
    df['tanggal'] = pd.to_datetime(df['tanggal']).dt.date
    return make_features(df)

def preprocess_csv_per_kategori(df_raw, test_size, n_estimators):
    col_map = detect_columns(df_raw)
    if 'tanggal' not in col_map or 'jumlah_produk' not in col_map:
        return None

    df = df_raw.copy()
    df['_tgl'] = pd.to_datetime(df[col_map['tanggal']], format='%d-%m-%Y %H:%M:%S', errors='coerce')
    df = df.dropna(subset=['_tgl'])
    df['tanggal'] = df['_tgl'].dt.date
    df[col_map['jumlah_produk']] = df[col_map['jumlah_produk']].apply(clean_number)

    hasil = {}
    kat_col = col_map.get('kategori')

    for label, keyword in [('makanan', 'makanan'), ('minuman', 'minuman')]:
        if kat_col:
            mask = df[kat_col].astype(str).str.lower().str.contains(keyword, na=False)
            sub  = df[mask]
        else:
            sub = df  # fallback jika tidak ada kolom kategori

        daily_kat = sub.groupby('tanggal').agg(
            jumlah_terjual=(col_map['jumlah_produk'], 'sum')
        ).reset_index()

        if len(daily_kat) >= 10:
            daily_feat = make_features(daily_kat)
            if len(daily_feat) >= 10:
                res = train_eval(daily_feat, test_size, n_estimators)
                res['kategori'] = label.capitalize()
                hasil[label]    = res

    return hasil if hasil else None

# ─────────────── ROUTES ───────────────
@app.route('/')
def dashboard():
    conn = get_db()
    total   = conn.execute('SELECT COUNT(*) as c FROM penjualan').fetchone()['c']
    makanan = conn.execute("SELECT COUNT(*) as c FROM penjualan WHERE LOWER(kategori)='makanan'").fetchone()['c']
    minuman = conn.execute("SELECT COUNT(*) as c FROM penjualan WHERE LOWER(kategori)='minuman'").fetchone()['c']
    rows_m  = conn.execute("SELECT tanggal, SUM(jumlah) as total FROM penjualan WHERE LOWER(kategori)='makanan' GROUP BY tanggal ORDER BY tanggal").fetchall()
    rows_n  = conn.execute("SELECT tanggal, SUM(jumlah) as total FROM penjualan WHERE LOWER(kategori)='minuman' GROUP BY tanggal ORDER BY tanggal").fetchall()
    conn.close()

    last_result = None
    try:
        conn2 = get_db()
        row = conn2.execute('SELECT result_json FROM prediksi_hasil ORDER BY id DESC LIMIT 1').fetchone()
        conn2.close()
        if row:
            last_result = json.loads(row['result_json'])
    except:
        pass

    return render_template('dashboard.html',
        total=total, makanan=makanan, minuman=minuman,
        chart_makanan=json.dumps([{'tanggal': r['tanggal'], 'total': r['total']} for r in rows_m]),
        chart_minuman=json.dumps([{'tanggal': r['tanggal'], 'total': r['total']} for r in rows_n]),
        last_result=last_result,
        produk_makanan=PRODUK_MAKANAN, produk_minuman=PRODUK_MINUMAN,
        active='dashboard'
    )

@app.route('/input-data')
def input_data():
    conn  = get_db()
    rows  = conn.execute('SELECT * FROM penjualan ORDER BY tanggal DESC LIMIT 100').fetchall()
    total = conn.execute('SELECT COUNT(*) as c FROM penjualan').fetchone()['c']
    mak   = conn.execute("SELECT COUNT(*) as c FROM penjualan WHERE LOWER(kategori)='makanan'").fetchone()['c']
    minn  = conn.execute("SELECT COUNT(*) as c FROM penjualan WHERE LOWER(kategori)='minuman'").fetchone()['c']
    conn.close()
    return render_template('input_data.html',
        rows=rows, total=total, makanan=mak, minuman=minn,
        today=date.today().strftime('%Y-%m-%d'),
        produk_makanan=PRODUK_MAKANAN, produk_minuman=PRODUK_MINUMAN,
        active='input'
    )

@app.route('/save-data', methods=['POST'])
def save_data():
    tanggal  = request.form.get('tanggal')
    produk   = request.form.get('produk')
    kategori = request.form.get('kategori', 'Makanan')
    jumlah   = int(request.form.get('jumlah', 0) or 0)
    harga    = int(request.form.get('harga', 0) or 0)
    conn = get_db()
    conn.execute('INSERT INTO penjualan (tanggal,produk,kategori,jumlah,harga) VALUES (?,?,?,?,?)',
                 (tanggal, produk, kategori, jumlah, harga))
    conn.commit(); conn.close()
    return redirect(url_for('input_data'))

@app.route('/reset-data', methods=['POST'])
def reset_data():
    try:
        conn = get_db()
        conn.execute('DELETE FROM penjualan')
        conn.execute('DELETE FROM prediksi_hasil')
        conn.commit()
        conn.close()
        flash('Semua data berhasil direset.', 'success')
    except Exception as e:
        flash(f'Gagal mereset data: {e}', 'error')
    return redirect(url_for('input_data'))

@app.route('/upload-csv', methods=['POST'])
def upload_csv():
    f = request.files.get('csv_file')
    if not f or f.filename == '':
        return redirect(url_for('input_data'))
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], f.filename)
    f.save(filepath)
    try:
        df      = pd.read_csv(filepath, low_memory=False)
        col_map = detect_columns(df)
        if 'tanggal' in col_map and 'jumlah_produk' in col_map:
            df['_tgl'] = pd.to_datetime(df[col_map['tanggal']], format='%d-%m-%Y %H:%M:%S', errors='coerce')
            df = df.dropna(subset=['_tgl'])
            df['_tgl_str'] = df['_tgl'].dt.strftime('%Y-%m-%d')
            df[col_map['jumlah_produk']] = df[col_map['jumlah_produk']].apply(clean_number)
            nama_col = col_map.get('nama_produk')
            kat_col  = col_map.get('kategori')
            hrg_col  = col_map.get('harga_produk', col_map.get('penjualan_kotor'))
            conn = get_db()
            for _, row in df.iterrows():
                produk   = str(row[nama_col])[:60]  if nama_col and pd.notna(row[nama_col]) else 'Import'
                kat_raw  = str(row[kat_col]).strip() if kat_col  and pd.notna(row[kat_col])  else 'Makanan'
                kategori = 'Minuman' if 'minuman' in kat_raw.lower() else 'Makanan'
                harga    = clean_number(row[hrg_col]) if hrg_col and pd.notna(row[hrg_col]) else 0
                jumlah   = int(row[col_map['jumlah_produk']])
                conn.execute('INSERT INTO penjualan (tanggal,produk,kategori,jumlah,harga) VALUES (?,?,?,?,?)',
                             (row['_tgl_str'], produk, kategori, jumlah, harga))
            conn.commit(); conn.close()
    except Exception as e:
        print(f'Upload error: {e}')
    return redirect(url_for('input_data'))

@app.route('/prediksi')
def prediksi():
    return render_template('prediksi.html', active='prediksi')

@app.route('/run-prediksi', methods=['POST'])
def run_prediksi():
    test_size    = float(request.form.get('test_size', 0.2))
    n_estimators = int(request.form.get('n_estimators', 100))
    source       = request.form.get('source', 'db')

    hasil = {}

    # Dari CSV
    if source == 'csv' and 'csv_file' in request.files:
        f = request.files['csv_file']
        if f.filename:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], f.filename)
            f.save(filepath)
            df_raw  = pd.read_csv(filepath, low_memory=False)
            hasil_csv = preprocess_csv_per_kategori(df_raw, test_size, n_estimators)
            if hasil_csv:
                hasil.update(hasil_csv)

    # Dari DB (fallback atau tambahan)
    for kat in ['makanan', 'minuman']:
        if kat not in hasil:
            daily = build_daily_from_db(kat)
            if daily is not None:
                res = train_eval(daily, test_size, n_estimators)
                res['kategori'] = kat.capitalize()
                hasil[kat]      = res

    if not hasil:
        return render_template('prediksi.html',
            error='Data tidak cukup. Minimal 10 hari data per kategori diperlukan. '
                  'Pastikan kolom Kategori berisi "Makanan" dan "Minuman".',
            active='prediksi')

    try:
        conn = get_db()
        conn.execute('INSERT INTO prediksi_hasil (created_at,result_json) VALUES (?,?)',
                     (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), json.dumps(hasil)))
        conn.commit(); conn.close()
    except Exception as e:
        print(f'Simpan error: {e}')

    session['prediksi_result'] = hasil
    return redirect(url_for('hasil'))

def _build_stok_per_item(buffer_pct=15, n_estimators=100, horizon=1):
    conn = get_db()
    produk_rows = conn.execute(
        'SELECT DISTINCT produk, kategori FROM penjualan ORDER BY kategori, produk'
    ).fetchall()
    conn.close()
    stok = {'makanan': [], 'minuman': []}
    for row in produk_rows:
        daily_raw, kat = build_daily_per_produk_from_db(row['produk'])
        if daily_raw is None:
            continue
        item = predict_item_horizon(daily_raw, horizon=horizon,
                                    n_estimators=n_estimators, buffer_pct=buffer_pct)
        item['produk']   = row['produk']
        item['kategori'] = row['kategori']
        kat_key = 'minuman' if 'minuman' in str(row['kategori']).lower() else 'makanan'
        stok[kat_key].append(item)
    for k in stok:
        stok[k].sort(key=lambda x: x['stok_anjuran'], reverse=True)
    return stok


@app.route('/hasil')
def hasil():
    result = session.get('prediksi_result')
    if not result:
        try:
            conn = get_db()
            row  = conn.execute('SELECT result_json FROM prediksi_hasil ORDER BY id DESC LIMIT 1').fetchone()
            conn.close()
            if row:
                result = json.loads(row['result_json'])
        except:
            pass
    if not result:
        return redirect(url_for('prediksi'))

    stok_items = _build_stok_per_item()
    next_date  = None
    all_items  = stok_items['makanan'] + stok_items['minuman']
    if all_items:
        next_date = all_items[0]['next_date']

    return render_template('hasil.html',
        result=result,
        result_json=json.dumps(result),
        stok_items=stok_items,
        next_date=next_date,
        active='hasil'
    )

@app.route('/export-hasil')
def export_hasil():
    result = session.get('prediksi_result')
    if not result:
        try:
            conn = get_db()
            row  = conn.execute('SELECT result_json FROM prediksi_hasil ORDER BY id DESC LIMIT 1').fetchone()
            conn.close()
            if row:
                result = json.loads(row['result_json'])
        except:
            pass
    if not result:
        return redirect(url_for('hasil'))

    hfill = PatternFill(start_color='2D6A2D', end_color='2D6A2D', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)

    def style_hdr(ws):
        for cell in ws[1]:
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal='center')

    wb    = openpyxl.Workbook()
    first = True
    for kat, res in result.items():
        label = res.get('kategori', kat.capitalize())
        ws = wb.active if first else wb.create_sheet()
        ws.title = f'Prediksi {label}'
        first = False
        ws.append(['No', 'Aktual', 'Prediksi LR', 'Prediksi RF'])
        style_hdr(ws)
        for i, (a, lr, rf) in enumerate(zip(res['y_test'], res['y_pred_lr'], res['y_pred_rf']), 1):
            ws.append([i, a, lr, rf])

        ws2 = wb.create_sheet(f'Evaluasi {label}')
        ws2.append(['Model', 'MAE', 'RMSE', 'Model Terbaik'])
        style_hdr(ws2)
        ws2.append(['Linear Regression', res['mae_lr'], res['rmse_lr'],
                    'Ya' if res['best_model'] == 'Linear Regression' else ''])
        ws2.append(['Random Forest', res['mae_rf'], res['rmse_rf'],
                    'Ya' if res['best_model'] == 'Random Forest' else ''])

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name='Hasil_Prediksi_Makanan_Minuman.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export-data')
def export_data():
    conn = get_db()
    rows = conn.execute('SELECT * FROM penjualan ORDER BY tanggal, kategori').fetchall()
    conn.close()
    df  = pd.DataFrame([dict(r) for r in rows])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df.to_excel(w, index=False, sheet_name='Data Penjualan')
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name='Data_Penjualan_Ayam_Serayu.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/stok-besok')
def stok_besok():
    buffer_pct   = int(request.args.get('buffer', 15))
    n_estimators = int(request.args.get('n_estimators', 100))
    horizon      = int(request.args.get('horizon', 1))
    if horizon not in (1, 7, 30):
        horizon = 1

    hasil = _build_stok_per_item(buffer_pct=buffer_pct,
                                  n_estimators=n_estimators,
                                  horizon=horizon)

    all_items = hasil['makanan'] + hasil['minuman']
    last_date = all_items[0]['last_date'] if all_items else None
    next_date = all_items[0]['next_date'] if all_items else None
    end_date  = all_items[0]['end_date']  if all_items else None

    total_stok_makanan = sum(i['stok_anjuran'] for i in hasil['makanan'])
    total_stok_minuman = sum(i['stok_anjuran'] for i in hasil['minuman'])

    horizon_label = {1: 'Besok', 7: '1 Minggu', 30: '1 Bulan'}[horizon]

    return render_template('stok_besok.html',
        hasil=hasil,
        buffer_pct=buffer_pct,
        n_estimators=n_estimators,
        horizon=horizon,
        horizon_label=horizon_label,
        last_date=last_date,
        next_date=next_date,
        end_date=end_date,
        total_stok_makanan=total_stok_makanan,
        total_stok_minuman=total_stok_minuman,
        active='stok'
    )


@app.route('/export-stok-besok')
def export_stok_besok():
    buffer_pct   = int(request.args.get('buffer', 15))
    n_estimators = int(request.args.get('n_estimators', 100))
    horizon      = int(request.args.get('horizon', 1))
    if horizon not in (1, 7, 30):
        horizon = 1

    conn = get_db()
    produk_rows = conn.execute(
        'SELECT DISTINCT produk, kategori FROM penjualan ORDER BY kategori, produk'
    ).fetchall()
    conn.close()

    stok = _build_stok_per_item(buffer_pct=buffer_pct,
                                n_estimators=n_estimators,
                                horizon=horizon)
    semua = stok['makanan'] + stok['minuman']
    semua.sort(key=lambda x: (x['kategori'], -x['stok_anjuran']))

    horizon_label = {1: 'Besok', 7: '1 Minggu', 30: '1 Bulan'}[horizon]
    pred_col = 'Prediksi Besok' if horizon == 1 else f'Total Prediksi ({horizon} Hari)'
    stok_col = f'Stok Anjuran (+{buffer_pct}%)'

    hfill = PatternFill(start_color='2D6A2D', end_color='2D6A2D', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True)

    def style_hdr(ws):
        for cell in ws[1]:
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal='center')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Stok {horizon_label} – Semua Item'
    ws.append(['No', 'Nama Produk', 'Kategori', 'Terjual Terakhir',
               'Rata-rata 7 Hari', 'Prediksi LR', 'Prediksi RF',
               pred_col, 'Avg/Hari', stok_col, 'Metode',
               'Data Terakhir', 'Mulai', 'Sampai', 'Jumlah Data'])
    style_hdr(ws)
    for i, item in enumerate(semua, 1):
        ws.append([
            i, item['produk'], item['kategori'],
            item['terjual_terakhir'], item['rata_7hari'],
            item['pred_lr'], item['pred_rf'],
            item['prediksi'], item['avg_per_hari'], item['stok_anjuran'], item['metode'],
            item['last_date'], item['next_date'], item['end_date'], item['total_data'],
        ])

    for kat_label_xl, kat_key in [('Makanan', 'makanan'), ('Minuman', 'minuman')]:
        ws2 = wb.create_sheet(f'Stok {kat_label_xl}')
        ws2.append(['No', 'Nama Produk', 'Terjual Terakhir',
                    'Rata-rata 7 Hari', 'Prediksi LR', 'Prediksi RF',
                    pred_col, 'Avg/Hari', stok_col, 'Metode',
                    'Data Terakhir', 'Mulai', 'Sampai'])
        style_hdr(ws2)
        filtered = [x for x in semua if kat_key in str(x['kategori']).lower()]
        for i, item in enumerate(filtered, 1):
            ws2.append([
                i, item['produk'], item['terjual_terakhir'], item['rata_7hari'],
                item['pred_lr'], item['pred_rf'], item['prediksi'],
                item['avg_per_hari'], item['stok_anjuran'], item['metode'],
                item['last_date'], item['next_date'], item['end_date'],
            ])

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    end_str = semua[0]['end_date'].replace('-', '') if semua else 'tgl'
    return send_file(buf, as_attachment=True,
        download_name=f'Stok_{horizon_label.replace(" ","")}_{end_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
